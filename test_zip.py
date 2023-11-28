import os
from zipfile import ZipFile
from xlrd import open_workbook
from openpyxl import load_workbook
from pypdf import PdfReader


file_dir = os.path.dirname(os.path.abspath(__file__))
res_dir = os.path.join(file_dir, 'resourses')
zip_tmp_dir = os.path.join(file_dir, 'tmp')
zip_dir = os.listdir(res_dir)
zip_path = os.path.join(zip_tmp_dir, 'test.zip')

if not os.path.exists(zip_tmp_dir):
    os.mkdir(zip_tmp_dir)

with ZipFile(zip_path, 'w') as zf:
    for file in zip_dir:
        add_file = os.path.join(res_dir, file)
        zf.write(add_file, file)


def test_find_files_in_archive():
    with ZipFile(zip_path, 'r') as zf:
        file_list = zf.namelist()
        assert 'file_example_XLS_10.xls' in file_list, 'Файл "file_example_XLS_10.xls" не найден'
        assert 'file_example_XLSX_50.xlsx' in file_list, 'Файл "file_example_XLSX_50.xlsx" не найден'
        assert 'Python_for_example.pdf' in file_list, 'Файл "Python_for_example.pdf" не найден'
        assert 'text_file.txt' in file_list, 'Файл "text_file.txt" не найден'


def test_correct_size_file_xlsx():
    with ZipFile(zip_path, 'r') as zf:
        xlsx_file = zf.getinfo('file_example_XLSX_50.xlsx')
        assert xlsx_file.file_size == 7360


def test_correct_size_file_xls():
    with ZipFile(zip_path, 'r') as zf:
        xls_file = zf.getinfo('file_example_XLS_10.xls')
        assert xls_file.file_size == 8704


def test_correct_size_file_pdf():
    with ZipFile(zip_path, 'r') as zf:
        pdf_file = zf.getinfo('Python_for_example.pdf')
        assert pdf_file.file_size == 10923806


def test_correct_size_file_txt():
    with ZipFile(zip_path, 'r') as zf:
        txt_file = zf.getinfo('text_file.txt')
        assert txt_file.file_size == 10


def test_correct_file_xls_value():
    with ZipFile(zip_path, 'r') as zf:
        with zf.open('file_example_XLS_10.xls', 'r') as xls_file:
            book = open_workbook(file_contents=(xls_file.read()))
            assert book.nsheets == 1
            assert book.sheet_names() == ['Sheet1']
            assert book.sheet_by_index(0).cell_value(7, 2) == "Hurn"


def test_correct_file_xlsx_value():
    with ZipFile(zip_path, 'r') as zf:
        with zf.open('file_example_XLSX_50.xlsx', 'r') as xlsx_file:
            book = load_workbook(xlsx_file)
            sheet = book.active
            assert book.sheetnames == ['Sheet1']
            assert sheet['B45'].value == "Willodean"


def test_correct_file_pdf_value():
    with ZipFile(zip_path, 'r') as zf:
        with zf.open('Python_for_example.pdf', 'r') as pdf_file:
            file = PdfReader(pdf_file)
            count_pages = len(file.pages)
            assert count_pages == 193
            assert "Никола Лейси" in file.pages[2].extract_text()


def test_correct_file_txt_value():
    with ZipFile(zip_path, 'r') as zf:
        with zf.open('text_file.txt', 'r') as txt_file:
            txt_file = txt_file.readlines()
            assert len(txt_file) == 2
            assert txt_file[1].decode() == 'text'